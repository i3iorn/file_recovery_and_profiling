import csv
import json
from pathlib import Path
from xml.etree import ElementTree as ET
import openpyxl


def convert_decorator(func):
    def wrapper(*args, **kwargs):
        if not isinstance(args[0], BaseConverter):
            raise TypeError(f"First argument must be an instance of {BaseConverter.__name__}")
        self = args[0]

        if kwargs.get('preserve_extension', False):
            kwargs['to_extension'] = self.input_file.suffix

        out_path = self.input_file.with_suffix(kwargs.get('to_extension', '.txt'))

        func(out_path=out_path, *args, **kwargs)
        return out_path

    return wrapper


def handle_conversion_errors(func):
    def wrapper(self, out_path, *args, **kwargs):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_path}")

        try:
            return func(self, out_path, *args, **kwargs)
        except Exception as e:
            raise RuntimeError(f"An error occurred during conversion: {str(e)}")

    return wrapper


class BaseConverter:
    def __init__(self, input_path: str, preserve_extension: bool = False, preserve_input: bool = False):
        self.input_path = input_path
        self.input_file = Path(input_path)
        self.__preserve_extension = preserve_extension
        self.__preserve_input = preserve_input

    @convert_decorator
    @handle_conversion_errors
    def convert(self, out_path: Path) -> Path:
        raise NotImplementedError


class CsvToXmlConverter(BaseConverter):
    def __init__(self, input_path: str, root_element: str, row_element: str, **kwargs):
        super().__init__(input_path, **kwargs)
        self.root_element = root_element
        self.row_element = row_element

    @convert_decorator
    @handle_conversion_errors
    def convert(self, out_path: Path) -> Path:
        data = []
        with open(self.input_path, 'r', encoding='utf-8') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                data.append(row)

        root = ET.Element(self.root_element)
        for item in data:
            row = ET.SubElement(root, self.row_element)
            for key, value in item.items():
                ET.SubElement(row, key).text = value

        tree = ET.ElementTree(root)
        tree.write(out_path, encoding='utf-8', xml_declaration=True)

        return out_path


class CsvToJsonConverter(BaseConverter):
    @convert_decorator
    @handle_conversion_errors
    def convert(
            self,
            out_path: Path,
    ) -> Path:
        data = []
        with open(self.input_path, 'r', encoding='utf-8') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                data.append(row)

        with open(out_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, indent=4)

        return out_path


class XlsxToCsvConverter(BaseConverter):
    @convert_decorator
    @handle_conversion_errors
    def convert(
            self,
            out_path: Path,
    ) -> Path:
        workbook = openpyxl.load_workbook(self.input_path)
        sheet = workbook.active

        with open(out_path, 'w', newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in sheet.iter_rows():
                csv_writer.writerow([cell.value for cell in row])
        return out_path


class CsvToXlsxConverter(BaseConverter):
    @convert_decorator
    @handle_conversion_errors
    def convert(
            self,
            out_path: Path,
    ) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(self.input_path, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)

        wb.save(out_path)
        return out_path


class PdfToCsvConverter(BaseConverter):
    @convert_decorator
    @handle_conversion_errors
    def convert(
            self,
            out_path: Path,
    ) -> Path:
        raise NotImplementedError


class RtfToCsvConverter(BaseConverter):
    @convert_decorator
    @handle_conversion_errors
    def convert(
            self,
            out_path: Path,
    ) -> Path:
        raise NotImplementedError


class DocxToCsvConverter(BaseConverter):
    @convert_decorator
    @handle_conversion_errors
    def convert(
            self,
            out_path: Path,
    ) -> Path:
        raise NotImplementedError


class DocToCsvConverter(BaseConverter):
    @convert_decorator
    @handle_conversion_errors
    def convert(
            self,
            out_path: Path,
    ) -> Path:
        raise NotImplementedError
